import os
from openpyxl import Workbook, load_workbook
from datetime import datetime


# Where our budget file lives
EXCEL_FILE = "output/spendsnap_budget.xlsx"

# The column headers for our Transactions sheet, in order
HEADERS = [
    "receipt_id",
    "date",
    "merchant_name",
    "category",
    "total_amount",
    "tax_amount",
    "currency",
    "payment_method",
    "items",
    "is_valid",
    "flags",
    "image_path",
    "processed_at"
]


def create_new_workbook() -> Workbook:
    """
    Creates a brand new Excel workbook with the Transactions sheet
    and the correct headers in the first row.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Transactions"

    # Write the headers into row 1
    sheet.append(HEADERS)

    return workbook


def write_to_excel(validated_data: dict) -> dict:
    """
    Takes the validated receipt data and appends it as a new row
    in the Transactions sheet of our Excel budget file.
    """

    # Step 1: Don't write invalid receipts to Excel
    if not validated_data.get("is_valid"):
        return {
            "success": False,
            "reason": "Receipt is not valid, skipped writing to Excel",
            "flags": validated_data.get("flags", [])
        }

    # Step 2: Make sure the output folder exists
    os.makedirs("output", exist_ok=True)

    # Step 3: Open the existing workbook, or create a new one
    if os.path.exists(EXCEL_FILE):
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook["Transactions"]
    else:
        workbook = create_new_workbook()
        sheet = workbook["Transactions"]

    # Step 4: Build the row of data, matching the order of HEADERS
    receipt_id = f"RCP-{sheet.max_row:04d}"  # e.g. RCP-0001, RCP-0002

    items_list = validated_data.get("items", [])
    items_text = ", ".join(items_list) if items_list else ""

    flags_list = validated_data.get("flags", [])
    flags_text = ", ".join(flags_list) if flags_list else ""

    date_value = validated_data.get("date") or "UNKNOWN_DATE"



    row = [
        receipt_id,
        date_value,
        validated_data.get("merchant_name"),
        validated_data.get("category"),
        validated_data.get("total_amount"),
        validated_data.get("tax_amount"),
        validated_data.get("currency"),
        validated_data.get("payment_method"),
        items_text,
        validated_data.get("is_valid"),
        flags_text,
        validated_data.get("image_path"),
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ]

    # Step 5: Append the row to the sheet
    sheet.append(row)

    # Step 6: Save the workbook back to disk
    workbook.save(EXCEL_FILE)

    return {
        "success": True,
        "receipt_id": receipt_id,
        "file_path": EXCEL_FILE
    }


# Run this block only when testing this file directly
if __name__ == "__main__":
    sample_validated_receipt = {
        "is_receipt": True,
        "merchant_name": "Walmart",
        "date": "2020-09-21",
        "total_amount": 46.42,
        "tax_amount": 3.5,
        "currency": "USD",
        "payment_method": "Card",
        "items": ["Yogurt", "Bananas"],
        "category": "Shopping",
        "confidence": 0.95,
        "notes": None,
        "image_path": "receipts/test.jpg",
        "is_valid": True,
        "is_duplicate": False,
        "flags": []
    }

    print("Writing sample receipt to Excel...\n")
    result = write_to_excel(sample_validated_receipt)

    import json
    print(json.dumps(result, indent=2))

